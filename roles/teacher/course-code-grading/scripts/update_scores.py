#!/usr/bin/env python3
"""
课程设计代码评分 — 分数回写工具

功能：
  --write   将评分结果写入原 Excel 文件的分数和备注列
  --report  生成评分统计报告（JSON）

用法：
  python update_scores.py --write --file "学生信息表.xlsx" --scores '[...]'
  python update_scores.py --report --file "学生信息表.xlsx"

--scores 参数格式（JSON 字符串或 JSON 文件路径）：
  [
    {"学号": "2024001", "分数": 85, "备注": "功能完整，代码规范"},
    {"学号": "2024002", "分数": 92, "备注": "实现优秀，有额外功能拓展"}
  ]
"""

import argparse
import csv
import json
import os
import re
import sys
from collections import OrderedDict


def load_excel(filepath):
    """加载 Excel 文件，返回 (workbook, sheet, headers, rows)"""
    try:
        import openpyxl
    except ImportError:
        print("[错误] 需要安装 openpyxl: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    headers = [h.strip() if h else '' for h in headers]

    # 读取数据行（从第 2 行开始）
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        row_data = OrderedDict()
        for i, cell in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = cell.value
        # 跳过空行
        if any(v is not None and str(v).strip() for v in row_data.values()):
            rows.append((row[0].row, row_data))  # (行号, 数据)

    return wb, ws, headers, rows


def save_excel(wb, filepath):
    """保存 Excel 文件"""
    wb.save(filepath)
    print(f"[OK] 文件已保存: {filepath}")


def find_column_indices(headers):
    """查找关键列的索引"""
    indices = {}
    # 标准化表头映射（支持多种可能的列名）
    col_map = {
        "学号": ["学号", "学生学号", "id", "student_id", "studentid"],
        "班级": ["班级", "班别", "class", "class_name"],
        "姓名": ["姓名", "名字", "name", "student_name"],
        "代码仓库地址": ["代码仓库地址", "仓库地址", "git地址", "repo_url", "repository", "github地址"],
        "分数": ["分数", "成绩", "得分", "评分", "score", "total_score"],
        "备注": ["备注", "评语", "评价", "note", "comment", "remarks"],
    }

    for key, aliases in col_map.items():
        for alias in aliases:
            for i, h in enumerate(headers):
                if h.lower().strip() == alias.lower().strip():
                    indices[key] = i
                    break
            if key in indices:
                break

    return indices


def write_scores(filepath, scores_data):
    """
    将评分结果写入 Excel 文件。

    Args:
        filepath: Excel 文件路径
        scores_data: 评分数据列表
            [{"学号": "...", "分数": 85, "备注": "..."}, ...]
    """
    wb, ws, headers, rows = load_excel(filepath)
    indices = find_column_indices(headers)

    # 验证必要列
    if "学号" not in indices:
        print(f"[错误] 找不到'学号'列。当前表头: {headers}")
        return False

    if "分数" not in indices:
        print(f"[错误] 找不到'分数'列。当前表头: {headers}")
        return False

    # 构建学号 -> 数据的查找映射
    score_map = {}
    for item in scores_data:
        sid = str(item.get("学号", "")).strip()
        if sid:
            score_map[sid] = item

    # 逐行匹配写入
    updated_count = 0
    not_found = []

    for row_number, row_data in rows:
        student_id = str(row_data.get("学号", "") or "").strip()
        if not student_id:
            continue

        if student_id in score_map:
            score_item = score_map[student_id]

            # 写分数
            score_value = score_item.get("分数")
            if score_value is not None:
                col_idx = indices["分数"]
                cell = ws.cell(row=row_number, column=col_idx + 1)
                try:
                    cell.value = int(score_value)
                except (ValueError, TypeError):
                    cell.value = score_value

            # 写备注
            note_value = score_item.get("备注")
            if note_value is not None:
                col_idx = indices["备注"]
                cell = ws.cell(row=row_number, column=col_idx + 1)
                cell.value = note_value

            updated_count += 1
        else:
            not_found.append(student_id)

    # 保存文件
    save_excel(wb, filepath)

    print(f"  更新记录数: {updated_count}/{len(rows)}")

    if not_found:
        print(f"  未匹配的学号 ({len(not_found)}):")
        for sid in not_found:
            print(f"    - {sid}")

    return True


def generate_report(filepath):
    """
    生成评分统计报告（JSON 格式）。

    Args:
        filepath: Excel 文件路径
    """
    _, _, headers, rows = load_excel(filepath)
    indices = find_column_indices(headers)

    if "分数" not in indices:
        print(json.dumps({"error": "找不到'分数'列", "headers": headers}, ensure_ascii=False))
        return

    scores = []
    for _, row_data in rows:
        score_value = row_data.get("分数")
        if score_value is not None:
            try:
                scores.append(int(score_value))
            except (ValueError, TypeError):
                pass

    if not scores:
        print(json.dumps({"error": "无有效分数数据"}, ensure_ascii=False))
        return

    # 等级映射
    def get_grade(score):
        if score >= 90:
            return "优秀"
        elif score >= 80:
            return "良好"
        elif score >= 70:
            return "中等"
        elif score >= 60:
            return "及格"
        else:
            return "不及格"

    # 统计
    total = len(scores)
    avg_score = sum(scores) / total
    max_score = max(scores)
    min_score = min(scores)
    pass_count = sum(1 for s in scores if s >= 60)
    excellent_count = sum(1 for s in scores if s >= 90)

    grades = {}
    for s in scores:
        g = get_grade(s)
        grades[g] = grades.get(g, 0) + 1

    report = {
        "total_students": total,
        "average_score": round(avg_score, 1),
        "max_score": max_score,
        "min_score": min_score,
        "pass_rate": f"{pass_count / total * 100:.1f}%",
        "excellent_rate": f"{excellent_count / total * 100:.1f}%",
        "grade_distribution": grades,
        "score_list": sorted(scores, reverse=True),
    }

    print(json.dumps(report, ensure_ascii=False, indent=2))


def main():
    parser = argparse.ArgumentParser(
        description="课程设计代码评分 — 分数回写工具",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  %(prog)s --write --file "学生信息表.xlsx" --scores '[{"学号": "2024001", "分数": 85}]'
  %(prog)s --write --file "学生信息表.xlsx" --scores scores.json
  %(prog)s --report --file "学生信息表.xlsx"
        """
    )

    parser.add_argument('--file', required=True,
                        help='学生信息 Excel 文件路径')
    parser.add_argument('--write', action='store_true',
                        help='写入模式：将评分写入 Excel')
    parser.add_argument('--report', action='store_true',
                        help='报告模式：生成评分统计报告')
    parser.add_argument('--scores', metavar='JSON字符串或文件路径',
                        help='评分数据 (JSON 字符串或 .json 文件路径)')

    args = parser.parse_args()

    if not args.write and not args.report:
        parser.print_help()
        sys.exit(1)

    if args.report:
        generate_report(args.file)
        return

    if args.write:
        if not args.scores:
            print("[错误] --write 模式需要 --scores 参数")
            sys.exit(1)

        # 解析 scores 参数
        scores_data = None
        if os.path.isfile(args.scores):
            with open(args.scores, 'r', encoding='utf-8') as f:
                scores_data = json.load(f)
        else:
            try:
                scores_data = json.loads(args.scores)
            except json.JSONDecodeError as e:
                print(f"[错误] JSON 解析失败: {e}")
                sys.exit(1)

        if not isinstance(scores_data, list):
            print("[错误] scores 必须是一个 JSON 数组")
            sys.exit(1)

        write_scores(args.file, scores_data)


if __name__ == "__main__":
    main()
