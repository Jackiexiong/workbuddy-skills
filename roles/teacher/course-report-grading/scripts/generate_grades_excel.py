#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
课程设计/实验报告成绩汇总 Excel 生成脚本

用法:
    python generate_grades_excel.py --input grades.json --output 数据库系统原理_课设成绩汇总.xlsx

输入 JSON 格式见 references/excel-output-spec.md。
维度(dimensions)和等级(grades)在 JSON 中声明，脚本可适配不同课程的结构。

依赖: openpyxl (pip install openpyxl)
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[ERROR] 缺少依赖 openpyxl，请运行: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ── 样式常量 ──

HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
NORMAL_FONT = Font(name="微软雅黑", size=10)
COMMENT_FONT = Font(name="微软雅黑", size=9)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
SUMMARY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# 等级着色（五级制：优秀/良好/中等/及格/不及格，兼容优/良/中/差/不及格）
GRADE_FILLS = {
    "优秀": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "良好": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "中等": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "及格": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "不及格": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "优": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "良": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
    "中": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "差": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}
GRADE_FONTS = {
    "优秀": Font(name="微软雅黑", size=10, color="006100", bold=True),
    "良好": Font(name="微软雅黑", size=10, color="1F4E79", bold=True),
    "中等": Font(name="微软雅黑", size=10, color="7F6000", bold=True),
    "及格": Font(name="微软雅黑", size=10, color="833C00", bold=True),
    "不及格": Font(name="微软雅黑", size=10, color="9C0006", bold=True),
    "优": Font(name="微软雅黑", size=10, color="006100", bold=True),
    "良": Font(name="微软雅黑", size=10, color="1F4E79", bold=True),
    "中": Font(name="微软雅黑", size=10, color="7F6000", bold=True),
    "差": Font(name="微软雅黑", size=10, color="833C00", bold=True),
}

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# 默认维度（数据库课设：三大题 50/30/20）
DEFAULT_DIMENSIONS = [
    {"key": "q1_modeling", "label": "题目1-建模设计(50)", "max": 50},
    {"key": "q2_management", "label": "题目2-管理优化(30)", "max": 30},
    {"key": "q3_development", "label": "题目3-应用开发(20)", "max": 20},
]

# 默认等级（数据库课设五级制）
DEFAULT_GRADES = ["优秀", "良好", "中等", "及格", "不及格"]


def grade_from_score(total):
    """总分映射等级（五级制）"""
    if total >= 90:
        return "优秀"
    elif total >= 80:
        return "良好"
    elif total >= 70:
        return "中等"
    elif total >= 60:
        return "及格"
    else:
        return "不及格"


def write_detail_sheet(ws, data, dimensions):
    """写入成绩明细表"""
    headers = ["班级", "学号", "姓名"]
    headers += [d["label"] for d in dimensions]
    headers += ["总分", "等级", "综合评语"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"

    students = data.get("students", [])
    n_dims = len(dimensions)
    total_col = 4 + n_dims
    grade_col = total_col + 1
    comment_col = total_col + 2

    for row_idx, s in enumerate(students, 2):
        ws.cell(row=row_idx, column=1, value=s.get("class", "")).border = THIN_BORDER
        id_cell = ws.cell(row=row_idx, column=2, value=str(s.get("id", "")))
        id_cell.border = THIN_BORDER
        id_cell.number_format = "@"
        ws.cell(row=row_idx, column=3, value=s.get("name", "")).border = THIN_BORDER

        scores = s.get("scores", {})
        for col, dim in enumerate(dimensions, 4):
            cell = ws.cell(row=row_idx, column=col, value=scores.get(dim["key"], 0))
            cell.border = THIN_BORDER
            cell.alignment = CENTER
            cell.font = NORMAL_FONT

        total = s.get("total")
        if total is None:
            total = sum(scores.get(d["key"], 0) for d in dimensions)
        total_cell = ws.cell(row=row_idx, column=total_col, value=total)
        total_cell.border = THIN_BORDER
        total_cell.alignment = CENTER
        total_cell.font = NORMAL_FONT

        grade = s.get("grade") or grade_from_score(total)
        grade_cell = ws.cell(row=row_idx, column=grade_col, value=grade)
        grade_cell.border = THIN_BORDER
        grade_cell.alignment = CENTER
        if grade in GRADE_FILLS:
            grade_cell.fill = GRADE_FILLS[grade]
            grade_cell.font = GRADE_FONTS.get(grade, NORMAL_FONT)

        comment_cell = ws.cell(row=row_idx, column=comment_col, value=s.get("comment", ""))
        comment_cell.border = THIN_BORDER
        comment_cell.alignment = LEFT_WRAP
        comment_cell.font = COMMENT_FONT

        for col in range(1, grade_col + 1):
            ws.cell(row=row_idx, column=col).alignment = CENTER
        ws.cell(row=row_idx, column=comment_col).alignment = LEFT_WRAP

    # 统计行
    if students:
        summary_row = len(students) + 2
        ws.cell(row=summary_row, column=1, value="平均分").font = Font(
            name="微软雅黑", size=10, bold=True
        )
        ws.cell(row=summary_row, column=1).fill = SUMMARY_FILL
        for col in range(2, 4):
            ws.cell(row=summary_row, column=col).fill = SUMMARY_FILL

        for col, dim in enumerate(dimensions, 4):
            col_letter = get_column_letter(col)
            formula = f"=AVERAGE({col_letter}2:{col_letter}{summary_row - 1})"
            cell = ws.cell(row=summary_row, column=col, value=formula)
            cell.font = Font(name="微软雅黑", size=10, bold=True)
            cell.fill = SUMMARY_FILL
            cell.alignment = CENTER
            cell.number_format = "0.0"

        total_letter = get_column_letter(total_col)
        total_cell = ws.cell(
            row=summary_row,
            column=total_col,
            value=f"=AVERAGE({total_letter}2:{total_letter}{summary_row - 1})",
        )
        total_cell.font = Font(name="微软雅黑", size=10, bold=True)
        total_cell.fill = SUMMARY_FILL
        total_cell.alignment = CENTER
        total_cell.number_format = "0.0"

        for col in range(grade_col, comment_col + 1):
            ws.cell(row=summary_row, column=col).fill = SUMMARY_FILL

    col_widths = [12, 15, 12] + [16] * n_dims + [8, 8, 60]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_stats_sheet(ws, data, dimensions, grades):
    """写入统计摘要表"""
    students = data.get("students", [])
    total_count = len(students)
    detail_sheet_name = ws.parent.sheetnames[0]

    n_dims = len(dimensions)
    total_col = 4 + n_dims
    grade_col = total_col + 1
    total_letter = get_column_letter(total_col)
    grade_letter = get_column_letter(grade_col)

    stats: list[tuple[str, object]] = [("实评人数", total_count)]
    if total_count:
        stats.append(("平均分", f"=AVERAGE('{detail_sheet_name}'!{total_letter}2:{total_letter}{total_count + 1})"))
        stats.append(("最高分", f"=MAX('{detail_sheet_name}'!{total_letter}2:{total_letter}{total_count + 1})"))
        stats.append(("最低分", f"=MIN('{detail_sheet_name}'!{total_letter}2:{total_letter}{total_count + 1})"))
        stats.append(("及格率(≥60)", f'=COUNTIF(\'{detail_sheet_name}\'!{total_letter}2:{total_letter}{total_count + 1},">=60")/{total_count}'))
        stats.append(("优秀率(≥90)", f'=COUNTIF(\'{detail_sheet_name}\'!{total_letter}2:{total_letter}{total_count + 1},">=90")/{total_count}'))

    ws.cell(row=1, column=1, value="统计项").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=1).alignment = CENTER
    ws.cell(row=1, column=2, value="值").font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL
    ws.cell(row=1, column=2).alignment = CENTER

    for i, (label, value) in enumerate(stats, 2):
        ws.cell(row=i, column=1, value=label).font = NORMAL_FONT
        cell = ws.cell(row=i, column=2, value=value)
        cell.font = NORMAL_FONT
        if "率" in label:
            cell.number_format = "0.0%"
        elif "分" in label:
            cell.number_format = "0.0"

    grade_start = len(stats) + 3
    ws.cell(row=grade_start, column=1, value="等级分布").font = Font(
        name="微软雅黑", size=11, bold=True
    )
    ws.cell(row=grade_start + 1, column=1, value="等级").font = HEADER_FONT
    ws.cell(row=grade_start + 1, column=1).fill = HEADER_FILL
    ws.cell(row=grade_start + 1, column=2, value="人数").font = HEADER_FONT
    ws.cell(row=grade_start + 1, column=2).fill = HEADER_FILL
    ws.cell(row=grade_start + 1, column=3, value="占比").font = HEADER_FONT
    ws.cell(row=grade_start + 1, column=3).fill = HEADER_FILL

    for i, grade in enumerate(grades):
        r = grade_start + 2 + i
        ws.cell(row=r, column=1, value=grade).font = NORMAL_FONT
        if total_count:
            ws.cell(
                row=r,
                column=2,
                value=f'=COUNTIF(\'{detail_sheet_name}\'!{grade_letter}2:{grade_letter}{total_count + 1},"{grade}")',
            ).font = NORMAL_FONT
            ws.cell(
                row=r,
                column=3,
                value=f'=COUNTIF(\'{detail_sheet_name}\'!{grade_letter}2:{grade_letter}{total_count + 1},"{grade}")/{total_count}',
            ).number_format = "0.0%"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12


def write_anomalies_sheet(ws, data):
    """写入异常清单"""
    anomalies = data.get("anomalies", [])
    ws.cell(row=1, column=1, value="文件名").font = HEADER_FONT
    ws.cell(row=1, column=1).fill = HEADER_FILL
    ws.cell(row=1, column=2, value="异常原因").font = HEADER_FONT
    ws.cell(row=1, column=2).fill = HEADER_FILL

    for i, a in enumerate(anomalies, 2):
        ws.cell(row=i, column=1, value=a.get("filename", "")).font = NORMAL_FONT
        ws.cell(row=i, column=2, value=a.get("reason", "")).font = NORMAL_FONT

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 40


def generate_excel(data, output_path):
    """生成 Excel 文件"""
    dimensions = data.get("dimensions", DEFAULT_DIMENSIONS)
    grades = data.get("grades", DEFAULT_GRADES)

    wb = Workbook()

    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "成绩明细"
    write_detail_sheet(ws1, data, dimensions)

    ws2 = wb.create_sheet("统计摘要")
    write_stats_sheet(ws2, data, dimensions, grades)

    if data.get("anomalies"):
        ws3 = wb.create_sheet("异常清单")
        write_anomalies_sheet(ws3, data)

    wb.save(output_path)
    print(f"[OK] 成绩汇总表已生成: {output_path}")
    print(f"     实评人数: {len(data.get('students', []))}")
    print(f"     异常数: {len(data.get('anomalies', []))}")


def main():
    parser = argparse.ArgumentParser(description="课程设计成绩汇总 Excel 生成器")
    parser.add_argument("--input", "-i", required=True, help="评分结果 JSON 文件路径")
    parser.add_argument("--output", "-o", help="输出 Excel 文件路径（默认: {课程名}_成绩汇总.xlsx）")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"[ERROR] 输入文件不存在: {input_path}", file=sys.stderr)
        sys.exit(1)

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    course = data.get("course", "课程")
    semester = data.get("semester", "")
    default_name = f"{course}_课设成绩汇总_{semester}.xlsx" if semester else f"{course}_课设成绩汇总.xlsx"
    output_path = args.output or default_name

    generate_excel(data, output_path)


if __name__ == "__main__":
    main()
